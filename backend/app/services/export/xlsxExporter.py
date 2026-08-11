"""
XLSX Exporter service (S15 / S20).

Generates Excel workbooks for adjustment grids and rent rolls with **live Excel formulas intact**
(`=SUM(...)`, `=AVERAGE(...)`, `=D2*(1+H2)`), while sanitizing user-supplied textual strings to
prevent formula injection attacks.
"""

from __future__ import annotations

import os
from collections.abc import Sequence
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.utils.security import sanitize_excel_cell

NAVY_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
FORMULA_FONT = Font(name="Calibri", size=11, bold=True, color="000080")
THIN_BORDER = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)


def export_adjustment_grid_xlsx(
    comparables: Sequence[dict],
    subject_area: float | Decimal,
    output_path: str,
) -> str:
    """
    Build XLSX adjustment grid with live Excel formulas and formula injection sanitization.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adjustment Grid"

    headers = [
        "Comparable Address",
        "Raw Price (₹)",
        "Area (sqft)",
        "Raw Rate (₹/sqft)",
        "Location Adj (%)",
        "Size Adj (%)",
        "Age/Condition Adj (%)",
        "Total Adj (%)",
        "Adjusted Rate (₹/sqft)",
    ]

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    start_row = 2
    for idx, comp in enumerate(comparables):
        row = start_row + idx
        raw_addr = comp.get("address", f"Comparable {idx + 1}")
        addr = sanitize_excel_cell(raw_addr)

        price = float(comp.get("sale_price", 10000000))
        area = float(comp.get("area", 1000))
        loc_adj = float(comp.get("location_adj", 0.02))
        size_adj = float(comp.get("size_adj", -0.01))
        age_adj = float(comp.get("age_adj", 0.00))

        # Formulas
        raw_rate_formula = f"=B{row}/C{row}"
        total_adj_formula = f"=SUM(E{row}:G{row})"
        adj_rate_formula = f"=D{row}*(1+H{row})"

        ws.cell(row=row, column=1, value=addr)
        ws.cell(row=row, column=2, value=price)
        ws.cell(row=row, column=3, value=area)
        ws.cell(row=row, column=4, value=raw_rate_formula).font = FORMULA_FONT
        ws.cell(row=row, column=5, value=loc_adj)
        ws.cell(row=row, column=6, value=size_adj)
        ws.cell(row=row, column=7, value=age_adj)
        ws.cell(row=row, column=8, value=total_adj_formula).font = FORMULA_FONT
        ws.cell(row=row, column=9, value=adj_rate_formula).font = FORMULA_FONT

        for c in range(1, 10):
            ws.cell(row=row, column=c).border = THIN_BORDER

    # Summary rows with live formulas
    last_data_row = start_row + len(comparables) - 1
    
    avg_row = last_data_row + 2
    ws.cell(row=avg_row, column=1, value="Average Raw Rate").font = BOLD_FONT
    ws.cell(row=avg_row, column=4, value=f"=AVERAGE(D2:D{last_data_row})").font = FORMULA_FONT

    concluded_row = avg_row + 1
    ws.cell(row=concluded_row, column=1, value="Concluded Mean Adjusted Rate (₹/sqft)").font = BOLD_FONT
    ws.cell(row=concluded_row, column=9, value=f"=AVERAGE(I2:I{last_data_row})").font = FORMULA_FONT

    subject_val_row = concluded_row + 1
    ws.cell(row=subject_val_row, column=1, value=f"Concluded Subject Value (Area: {subject_area} sqft)").font = BOLD_FONT
    ws.cell(row=subject_val_row, column=9, value=f"=I{concluded_row}*{float(subject_area)}").font = FORMULA_FONT

    # Formatting column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(output_path)
    return output_path


def export_rent_roll_xlsx(
    units: Sequence[dict],
    output_path: str,
) -> str:
    """
    Build XLSX rent roll with live Excel formulas and formula injection sanitization.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rent Roll"

    headers = [
        "Tenant Name",
        "Unit / Area (sqft)",
        "Monthly Rent (₹)",
        "Annual Rent (₹)",
        "Security Deposit (₹)",
        "Escalation Rate (%)",
    ]

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    start_row = 2
    for idx, u in enumerate(units):
        row = start_row + idx
        raw_tenant = u.get("tenant_name", f"Tenant {idx + 1}")
        tenant = sanitize_excel_cell(raw_tenant)

        area = float(u.get("area", 1200))
        m_rent = float(u.get("monthly_rent", 150000))
        deposit = float(u.get("security_deposit", 900000))
        esc = float(u.get("escalation_rate", 0.05))

        annual_formula = f"=C{row}*12"

        ws.cell(row=row, column=1, value=tenant)
        ws.cell(row=row, column=2, value=area)
        ws.cell(row=row, column=3, value=m_rent)
        ws.cell(row=row, column=4, value=annual_formula).font = FORMULA_FONT
        ws.cell(row=row, column=5, value=deposit)
        ws.cell(row=row, column=6, value=esc)

        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER

    last_row = start_row + len(units) - 1
    total_row = last_row + 2

    ws.cell(row=total_row, column=1, value="PORTFOLIO TOTAL").font = BOLD_FONT
    ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{last_row})").font = FORMULA_FONT
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{last_row})").font = FORMULA_FONT
    ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{last_row})").font = FORMULA_FONT

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    wb.save(output_path)
    return output_path
