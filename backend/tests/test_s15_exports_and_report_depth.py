"""
Sprint 15 Multi-Format Exporters & Report Depth Tests.

Verifies:
1. Live Excel formulas intact in generated XLSX files (openpyxl formula inspection: =SUM, =AVERAGE, =B2*(1+C2)).
2. DOCX, PDF, and JSON exporters file generation.
3. IBBI / RICS reporting checklist sections coverage.
"""

from __future__ import annotations

import json
import os
from decimal import Decimal

import openpyxl

from app.services.export import docxExporter, jsonExporter, pdfExporter, xlsxExporter
from app.services.graph.prompts.valuationPrompts import IBBI_RICS_CHECKLIST_SECTIONS


def test_xlsx_exporter_has_live_formulas(tmp_path):
    """
    Verify XLSX adjustment grid and rent roll contain live Excel formulas (=SUM, =AVERAGE, etc.).
    """
    grid_path = str(tmp_path / "test_adjustment_grid.xlsx")
    rent_path = str(tmp_path / "test_rent_roll.xlsx")

    comps = [
        {"address": "Tower A Sector 62", "sale_price": 12000000, "area": 1500, "location_adj": 0.05, "size_adj": -0.02, "age_adj": 0.00},
        {"address": "Tower B Sector 62", "sale_price": 10500000, "area": 1400, "location_adj": 0.02, "size_adj": 0.00, "age_adj": -0.01},
    ]

    # 1. Export adjustment grid
    xlsxExporter.export_adjustment_grid_xlsx(comps, subject_area=Decimal("1500"), output_path=grid_path)
    wb_grid = openpyxl.load_workbook(grid_path, data_only=False)
    ws_grid = wb_grid["Adjustment Grid"]

    # Verify formula cells
    raw_rate_cell = ws_grid.cell(row=2, column=4).value
    assert isinstance(raw_rate_cell, str) and raw_rate_cell.startswith("=")
    assert raw_rate_cell == "=B2/C2"

    total_adj_cell = ws_grid.cell(row=2, column=8).value
    assert isinstance(total_adj_cell, str) and total_adj_cell.startswith("=")
    assert total_adj_cell == "=SUM(E2:G2)"

    adj_rate_cell = ws_grid.cell(row=2, column=9).value
    assert isinstance(adj_rate_cell, str) and adj_rate_cell.startswith("=")
    assert adj_rate_cell == "=D2*(1+H2)"

    avg_raw_cell = ws_grid.cell(row=5, column=4).value
    assert isinstance(avg_raw_cell, str) and "AVERAGE" in avg_raw_cell

    # 2. Export rent roll
    units = [
        {"tenant_name": "Acme Corp", "area": 2500, "monthly_rent": 250000, "security_deposit": 1500000, "escalation_rate": 0.05},
        {"tenant_name": "Beta LLC", "area": 1800, "monthly_rent": 180000, "security_deposit": 1080000, "escalation_rate": 0.05},
    ]
    xlsxExporter.export_rent_roll_xlsx(units, output_path=rent_path)
    wb_rent = openpyxl.load_workbook(rent_path, data_only=False)
    ws_rent = wb_rent["Rent Roll"]

    annual_rent_cell = ws_rent.cell(row=2, column=4).value
    assert isinstance(annual_rent_cell, str) and annual_rent_cell == "=C2*12"

    total_rent_cell = ws_rent.cell(row=5, column=3).value
    assert isinstance(total_rent_cell, str) and "SUM" in total_rent_cell


def test_pdf_docx_json_exporters(tmp_path):
    """Verify PDF, DOCX, and JSON exports generate valid files."""
    clause_plan = [
        {"heading": "Executive Summary", "type": "executive_summary", "content": "Market value is ₹ 2,50,00,000."},
        {"heading": "Property Description", "type": "property_description", "content": "Commercial office space."},
    ]

    # DOCX
    docx_path = docxExporter.export_valuation_docx(clause_plan, "valuation_report", "Client A", "Address A", {}, "job_exp_1")
    assert os.path.exists(docx_path) and os.path.getsize(docx_path) > 0

    # PDF
    pdf_path = pdfExporter.export_valuation_pdf(clause_plan, "valuation_report", "Client A", "Address A", "job_exp_1")
    assert os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0

    # JSON
    json_path = jsonExporter.export_valuation_json(clause_plan, "valuation_report", "Client A", "Address A", {"val": 25000000}, "job_exp_1")
    assert os.path.exists(json_path) and os.path.getsize(json_path) > 0

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    assert data["job_id"] == "job_exp_1"
    assert len(data["sections"]) == 2


def test_ibbi_rics_reporting_checklist_sections():
    """Verify IBBI/RICS checklist sections coverage."""
    assert len(IBBI_RICS_CHECKLIST_SECTIONS) == 9
    assert "scope_and_purpose" in IBBI_RICS_CHECKLIST_SECTIONS
    assert "basis_and_premise" in IBBI_RICS_CHECKLIST_SECTIONS
    assert "adjustment_grid" in IBBI_RICS_CHECKLIST_SECTIONS
    assert "valuer_declaration" in IBBI_RICS_CHECKLIST_SECTIONS
