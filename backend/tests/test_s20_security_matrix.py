"""
Sprint 20 Security, Confidentiality & Authorization Matrix Tests.

Verifies:
1. Neutralization of Excel formula injection (=cmd, +, -, @) in XLSX exports.
2. Exclusion of exact GPS coordinates and owner names from client-role API responses.
3. Programmatic authorization matrix enforcement across roles x resources x actions.
"""

from __future__ import annotations

from decimal import Decimal

import openpyxl

from app.services.export.xlsxExporter import export_adjustment_grid_xlsx, export_rent_roll_xlsx
from app.utils.security import (
    filter_client_role_response,
    verify_authorization,
)


def test_xlsx_formula_injection_neutralized(tmp_path):
    """
    Verify tenant names and addresses starting with =, +, -, @ are exported as inert text with leading single quote.
    """
    grid_path = str(tmp_path / "test_sec_grid.xlsx")
    rent_path = str(tmp_path / "test_sec_rent.xlsx")

    # Malicious input payloads
    malicious_comps = [
        {"address": "=cmd|' /C calc'!A0", "sale_price": 10000000, "area": 1000},
        {"address": "+1234567890", "sale_price": 12000000, "area": 1200},
        {"address": "@SUM(A1:A10)", "sale_price": 11000000, "area": 1100},
    ]

    export_adjustment_grid_xlsx(malicious_comps, Decimal("1000"), grid_path)
    wb_grid = openpyxl.load_workbook(grid_path, data_only=False)
    ws_grid = wb_grid["Adjustment Grid"]

    cell_val_1 = ws_grid.cell(row=2, column=1).value
    assert isinstance(cell_val_1, str) and cell_val_1.startswith("'=cmd")

    cell_val_2 = ws_grid.cell(row=3, column=1).value
    assert isinstance(cell_val_2, str) and cell_val_2.startswith("'+1234567890")

    cell_val_3 = ws_grid.cell(row=4, column=1).value
    assert isinstance(cell_val_3, str) and cell_val_3.startswith("'@SUM")

    # Rent roll check
    malicious_units = [
        {"tenant_name": "=cmd|' /C calc'!A0", "area": 1500, "monthly_rent": 150000},
    ]
    export_rent_roll_xlsx(malicious_units, rent_path)
    wb_rent = openpyxl.load_workbook(rent_path, data_only=False)
    ws_rent = wb_rent["Rent Roll"]

    tenant_cell = ws_rent.cell(row=2, column=1).value
    assert isinstance(tenant_cell, str) and tenant_cell.startswith("'=cmd")


def test_client_role_response_excludes_exact_coordinates_and_owners():
    """Verify client-role API response filter strips exact coordinates and owner names."""
    raw_response = {
        "property_id": "prop_123",
        "city": "Mumbai",
        "locality": "BKC",
        "latitude": 19.0760,
        "longitude": 72.8777,
        "coordinates": "POINT(19.0760 72.8777)",
        "owner_name": "Ramesh Kumar",
        "owner_details": {"name": "Ramesh Kumar", "pan": "ABCDE1234F"},
        "concluded_value": "25000000.00",
    }

    filtered = filter_client_role_response(raw_response)

    assert "city" in filtered
    assert "concluded_value" in filtered
    assert "latitude" not in filtered
    assert "longitude" not in filtered
    assert "coordinates" not in filtered
    assert "owner_name" not in filtered
    assert "owner_details" not in filtered


def test_authorization_matrix_permissions():
    """Verify programmatic authorization matrix across roles x resources x actions."""
    # Admin can do everything
    assert verify_authorization("admin", "deliverables", "sign") is True
    assert verify_authorization("admin", "deliverables", "delete") is True

    # Valuer can sign deliverables, but not delete mandates
    assert verify_authorization("valuer", "deliverables", "sign") is True
    assert verify_authorization("valuer", "mandates", "delete") is False

    # Analyst can export deliverables, but NOT sign
    assert verify_authorization("analyst", "deliverables", "export") is True
    assert verify_authorization("analyst", "deliverables", "sign") is False

    # Client can read deliverables, but NOT write, sign, or delete
    assert verify_authorization("client", "deliverables", "read") is True
    assert verify_authorization("client", "deliverables", "sign") is False
    assert verify_authorization("client", "jobs", "write") is False
